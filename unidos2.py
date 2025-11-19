# file: unidos2_limpio.py
import re
import unicodedata
import logging
import threading
from pathlib import Path
from typing import List, Tuple, Optional, Any, Dict

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pandas as pd
import pdfplumber
from unidecode import unidecode
from fuzzywuzzy import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ------------------------------
# Config / Constantes
# ------------------------------
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

CARPETA_DESCARGAS = Path.home() / "Downloads"

# Regex patterns
PATRON_GENERAL_NOMBRE = re.compile(r"A nombre de:\s*(.+?)\s*Estado:", re.IGNORECASE)
PATRON_REGISTRO_CIVIL = re.compile(r"Registro Civil,\s*(.*?)\s*tiene inscrito", re.IGNORECASE | re.DOTALL)
PATRON_MIGRACION_NOMBRE = re.compile(r"el migrante venezolano\s+([A-ZÁÉÍÓÚÑ ]{5,})\s+surtió", re.IGNORECASE | re.DOTALL)

PATRON_CEDULA_ADULTO = re.compile(r"Cédula de Ciudadanía:\s*([0-9]{1,3}(?:\.[0-9]{3})*\.[0-9]{3})", re.IGNORECASE)
PATRON_NUIP_MENOR = re.compile(r"Número Único de Identificación Personal\s+([0-9]+)", re.IGNORECASE)
PATRON_RUMV_PPT = re.compile(r"número de RUMV\s+([0-9]+)", re.IGNORECASE)

# Excel style fills
FILL_VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_AMARILLO = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_ROJO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_AZUL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

# Standardized thresholds and messages (recommended)
UMBRAL_EXACTA = 100
UMBRAL_ALTA = 90
UMBRAL_MEDIA = 70
UMBRAL_BAJA = 50

MSG_EXACTA = "✔ EXACTA"
MSG_ALTA = "⚠️ ALTA"
MSG_MEDIA = "⚠️ MEDIA"
MSG_BAJA = "⚠️ BAJA"
MSG_SIN = "❌ SIN COINCIDENCIA"
MSG_NO_EXTRAIDO = "NO EXTRAÍDO"
MSG_DOC_VACIO = "DOCUMENTO VACIO"

# ------------------------------
# UTILIDADES
# ------------------------------
def leer_pdf_completo(pdf_path: Path) -> str:
    """Leer todo el texto de un PDF en una sola pasada."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            paginas = [p.extract_text() for p in pdf.pages if p.extract_text()]
            texto = "\n".join(paginas)
            logging.debug("Leído PDF: %s (len=%d)", pdf_path.name, len(texto))
            return texto
    except Exception as e:
        logging.error("Error al abrir PDF %s: %s", pdf_path, e)
        return ""

def leer_columna_excel(path: str, index_col: int, skiprows: int = 6) -> List[str]:
    """Lee una columna de Excel comúnmente estructurada (skiprows por defecto 6)."""
    try:
        df = pd.read_excel(path, usecols=[index_col], skiprows=skiprows, header=None)
        valores = df.iloc[:, 0].dropna().astype(str).tolist()
        logging.debug("Leído Excel: %s columnas=%d items=%d", path, index_col, len(valores))
        return valores
    except Exception as e:
        logging.error("Error leyendo Excel %s: %s", path, e)
        raise

def normalizar_texto(texto: Any) -> str:
    """Normalize string: uppercase, remove accents, keep letters and spaces."""
    if texto is None:
        return ""
    s = str(texto).strip().upper()
    s = unicodedata.normalize("NFKD", s).encode("ASCII", "ignore").decode("utf-8")
    s = re.sub(r"[^A-Z\s]", "", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()

def normalizar_nombre_light(nombre: Any) -> str:
    """Light normalizer using unidecode, preserves ordering but uppercase."""
    if nombre is None:
        return ""
    return unidecode(str(nombre).strip().upper())

def normalizar_documento(documento: Any) -> str:
    """Keep only digits from document."""
    if documento is None:
        return ""
    return re.sub(r"[^0-9]", "", str(documento))

def invertir_nombre_si_correspondiente(nombre: str) -> str:
    partes = nombre.strip().split()
    if len(partes) >= 2:
        mitad = len(partes) // 2
        return " ".join(partes[mitad:] + partes[:mitad])
    return nombre

def evaluar_similitud(score: int) -> Tuple[str, PatternFill]:
    """Estándar de estados para puntuaciones fuzzy."""
    if score == UMBRAL_EXACTA:
        return MSG_EXACTA, FILL_VERDE
    if score >= UMBRAL_ALTA:
        return f"{MSG_ALTA} ({score}%)", FILL_AMARILLO
    if score >= UMBRAL_MEDIA:
        return f"{MSG_MEDIA} ({score}%)", FILL_AMARILLO
    if score >= UMBRAL_BAJA:
        return f"{MSG_BAJA} ({score}%)", FILL_AZUL
    return MSG_SIN, FILL_ROJO

# ------------------------------
# EXTRACCION (usa leer_pdf_completo)
# ------------------------------
def extraer_nombre_desde_texto(texto: str) -> Optional[str]:
    if not texto:
        return None
    match = PATRON_GENERAL_NOMBRE.search(texto)
    if match:
        return match.group(1).strip()
    match = PATRON_REGISTRO_CIVIL.search(texto)
    if match:
        return invertir_nombre_si_correspondiente(match.group(1).strip())
    match = PATRON_MIGRACION_NOMBRE.search(texto)
    if match:
        return match.group(1).strip()
    return None

def extraer_documento_desde_texto(texto: str) -> Tuple[Optional[str], Optional[str]]:
    if not texto:
        return None, None
    match = PATRON_CEDULA_ADULTO.search(texto)
    if match:
        return normalizar_documento(match.group(1)), "CEDULA_ADULTO"
    match = PATRON_NUIP_MENOR.search(texto)
    if match:
        return normalizar_documento(match.group(1)), "NUIP_MENOR"
    match = PATRON_RUMV_PPT.search(texto)
    if match:
        return normalizar_documento(match.group(1)), "RUMV_PPT"
    return None, None

# ------------------------------
# COMPARADORES
# ------------------------------
def comparar_nombres_fuzzy(nombres_pdf: List[str], nombres_excel: List[str]) -> List[Dict[str, Any]]:
    resultados = []
    # Normalize Excel names once (light)
    nombres_excel_norm = [normalizar_texto(n) for n in nombres_excel]

    for nombre_pdf in nombres_pdf:
        nombre_pdf_norm = normalizar_texto(nombre_pdf)
        mejor_match = None
        mejor_score = -1

        for i, nombre_excel in enumerate(nombres_excel):
            score = fuzz.token_sort_ratio(nombre_pdf_norm, nombres_excel_norm[i])
            if score > mejor_score:
                mejor_score = score
                mejor_match = nombre_excel

        estado, _ = evaluar_similitud(mejor_score if mejor_score >= 0 else 0)
        resultados.append({
            "Nombre PDF": nombre_pdf,
            "Mejor Coincidencia Excel": mejor_match or "—",
            "Porcentaje Similitud": mejor_score if mejor_score >= 0 else 0,
            "Estado": estado
        })

    # Excel -> PDFs check for missing certificates
    for nombre_excel, nombre_excel_norm in zip(nombres_excel, nombres_excel_norm):
        # if no PDF name reaches ALTA threshold, mark as missing certificate
        if not any(fuzz.token_sort_ratio(nombre_excel_norm, normalizar_texto(n)) >= UMBRAL_ALTA for n in nombres_pdf):
            resultados.append({
                "Nombre PDF": "—",
                "Mejor Coincidencia Excel": nombre_excel,
                "Porcentaje Similitud": 0,
                "Estado": "FALTA CERTIFICADO"
            })

    return resultados

def comparar_documentos_exactos(documentos_pdf: List[Tuple[str, str, str]], documentos_excel: List[str]) -> List[List[Any]]:
    resultados = []
    documentos_excel_copy = documentos_excel.copy()

    for numero_doc_pdf, tipo_doc, filename in documentos_pdf:
        if not numero_doc_pdf:
            resultados.append([filename, numero_doc_pdf or "", tipo_doc or "", "", 0, MSG_DOC_VACIO])
            continue

        if numero_doc_pdf in documentos_excel_copy:
            documentos_excel_copy.remove(numero_doc_pdf)
            resultados.append([filename, numero_doc_pdf, tipo_doc, numero_doc_pdf, 100, MSG_EXACTA])
            continue

        mejor_sim = 0
        mejor_match = ""
        for doc_excel in documentos_excel:
            sim = fuzz.ratio(numero_doc_pdf, doc_excel)
            if sim > mejor_sim:
                mejor_sim = sim
                mejor_match = doc_excel

        estado, _ = evaluar_similitud(mejor_sim)
        resultados.append([filename, numero_doc_pdf, tipo_doc, mejor_match, mejor_sim, estado])

    return resultados

# ------------------------------
# EXPORTADORES (reutilizables)
# ------------------------------
def _formatear_encabezados(ws):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

def _aplicar_color_por_estado(ws, estado_col_idx_map: Dict[int, int]):
    """
    estado_col_idx_map: mapping from columna_estado_index -> (start_col, end_col)
    e.g. {3: (1,4)} means column 3 has estado para columnas 1..4
    """
    for row_idx in range(2, ws.max_row + 1):
        # determine fill for each estado column and apply to its column span
        for estado_col, (start_col, end_col) in estado_col_idx_map.items():
            estado_val = str(ws.cell(row=row_idx, column=estado_col).value or "")
            if "EXACTA" in estado_val:
                fill = FILL_VERDE
            elif "ALTA" in estado_val or "MEDIA" in estado_val:
                fill = FILL_AMARILLO
            elif "BAJA" in estado_val:
                fill = FILL_AZUL
            else:
                fill = FILL_ROJO
            for c in range(start_col, end_col + 1):
                ws.cell(row=row_idx, column=c).fill = fill

def exportar_dataframe_con_formato(df: pd.DataFrame, ruta_salida: Path, estado_col_idx_map: Dict[int, Tuple[int,int]]):
    df.to_excel(ruta_salida, index=False)
    wb = load_workbook(ruta_salida)
    ws = wb.active
    _formatear_encabezados(ws)
    _aplicar_color_por_estado(ws, estado_col_idx_map)
    wb.save(ruta_salida)

def exportar_resultados_nombres(resultados: List[Dict[str, Any]]):
    df = pd.DataFrame(resultados, columns=["Nombre PDF", "Mejor Coincidencia Excel", "Porcentaje Similitud", "Estado"])
    ruta = CARPETA_DESCARGAS / "comparacion_nombres.xlsx"
    try:
        exportar_dataframe_con_formato(df, ruta, {4: (1, 4)})  # estado en col 4 -> aplica a A:D
        messagebox.showinfo("Éxito", f"✅ Comparación de nombres completada\n📁 Guardado en: {ruta}")
    except Exception as e:
        logging.exception("Error exportando nombres: %s", e)
        messagebox.showerror("Error", f"Error al exportar: {e}")

def exportar_resultados_documentos(resultados: List[List[Any]]):
    df = pd.DataFrame(resultados, columns=[
        "Archivo PDF", "Documento Extraído", "Tipo Documento",
        "Mejor Coincidencia Excel", "% Similitud", "Estado"
    ])
    ruta = CARPETA_DESCARGAS / "comparacion_documentos.xlsx"
    try:
        exportar_dataframe_con_formato(df, ruta, {6: (1, 6)})  # estado en col 6 -> aplica a A:F
        messagebox.showinfo("Éxito", f"✅ Comparación de documentos completada\n📁 Guardado en: {ruta}")
    except Exception as e:
        logging.exception("Error exportando documentos: %s", e)
        messagebox.showerror("Error", f"Error al exportar: {e}")

def exportar_resultados_completos(resultados: List[List[Any]]):
    ruta = CARPETA_DESCARGAS / "validacion_completa.xlsx"
    try:
        df = pd.DataFrame(resultados, columns=[
            "Archivo PDF", "Nombre Extraído", "Mejor Nombre Excel", "% Similitud Nombre", "Estado Nombre",
            "Documento Extraído", "Tipo Documento", "Mejor Documento Excel", "% Similitud Documento", "Estado Documento"
        ])
        with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Validación Completa", index=False)

            # Estadísticas sencillas
            total = len(resultados) or 1
            nombres_exactos = sum(1 for r in resultados if "EXACTA" in str(r[4]))
            nombres_similares = sum(1 for r in resultados if "ALTA" in str(r[4]) or "MEDIA" in str(r[4]))
            nombres_sin = sum(1 for r in resultados if "SIN COINCIDENCIA" in str(r[4]) or "NO EXTRAÍDO" in str(r[4]))

            docs_exactos = sum(1 for r in resultados if "EXACTA" in str(r[9]))
            docs_similares = sum(1 for r in resultados if "ALTA" in str(r[9]) or "MEDIA" in str(r[9]))
            docs_sin = sum(1 for r in resultados if "SIN COINCIDENCIA" in str(r[9]) or "NO EXTRAÍDO" in str(r[9]))

            stats = [
                ["Métrica", "Cantidad", "Porcentaje"],
                ["Total de archivos procesados", total, "100%"],
                [],
                ["VALIDACIÓN DE NOMBRES", "", ""],
                ["Coincidencias exactas", nombres_exactos, f"{(nombres_exactos/total)*100:.1f}%"],
                ["Similitudes parciales", nombres_similares, f"{(nombres_similares/total)*100:.1f}%"],
                ["Sin coincidencia", nombres_sin, f"{(nombres_sin/total)*100:.1f}%"],
                [],
                ["VALIDACIÓN DE DOCUMENTOS", "", ""],
                ["Coincidencias exactas", docs_exactos, f"{(docs_exactos/total)*100:.1f}%"],
                ["Similitudes parciales", docs_similares, f"{(docs_similares/total)*100:.1f}%"],
                ["Sin coincidencia", docs_sin, f"{(docs_sin/total)*100:.1f}%"],
            ]
            df_stats = pd.DataFrame(stats)
            df_stats.to_excel(writer, sheet_name="Estadísticas", index=False)

        # Formatear hoja principal
        wb = load_workbook(ruta)
        ws = wb["Validación Completa"]
        _formatear_encabezados(ws)
        # Estado Nombre en col 5 -> aplica A:E ; Estado Documento en col 10 -> aplica F:J
        _aplicar_color_por_estado(ws, {5: (1, 5), 10: (6, 10)})
        wb.save(ruta)
        messagebox.showinfo("Éxito", f"✅ Validación completa finalizada\n📁 Guardado en: {ruta}\n\n📊 Estadísticas incluidas")
    except Exception as e:
        logging.exception("Error exportando validación completa: %s", e)
        messagebox.showerror("Error", f"Error al exportar: {e}")

# ------------------------------
# FUNCIONES DE ALTO NIVEL USADAS POR LA GUI (mantienen comportamiento)
# ------------------------------
def procesar_extraer_nombres(carpeta: Path):
    nombres_extraidos = []
    for item in carpeta.iterdir():
        if item.is_file() and item.suffix.lower() == ".pdf":
            texto = leer_pdf_completo(item)
            nombre = extraer_nombre_desde_texto(texto)
            if nombre:
                nombres_extraidos.append([item.name, nombre])
    if nombres_extraidos:
        df = pd.DataFrame(nombres_extraidos, columns=["Archivo PDF", "Nombre Extraído"])
        ruta = CARPETA_DESCARGAS / "nombres_extraidos.xlsx"
        df.to_excel(ruta, index=False)
        return ruta, len(nombres_extraidos)
    return None, 0

def procesar_extraer_documentos(carpeta: Path):
    documentos = []
    for item in carpeta.iterdir():
        if item.is_file() and item.suffix.lower() == ".pdf":
            texto = leer_pdf_completo(item)
            num, tipo = extraer_documento_desde_texto(texto)
            if num and tipo:
                documentos.append([item.name, num, tipo])
    if documentos:
        df = pd.DataFrame(documentos, columns=["Archivo PDF", "Número Documento", "Tipo Documento"])
        ruta = CARPETA_DESCARGAS / "documentos_extraidos.xlsx"
        df.to_excel(ruta, index=False)
        return ruta, len(documentos)
    return None, 0

def procesar_comparar_nombres(carpeta: Path, excel_path: str):
    nombres_pdf = []
    # Extraer nombres de PDFs (una sola pasada por archivo)
    for item in Path(carpeta).iterdir():
        if item.is_file() and item.suffix.lower() == ".pdf":
            texto = leer_pdf_completo(item)
            nombre = extraer_nombre_desde_texto(texto)
            if nombre:
                nombres_pdf.append(nombre)
    nombres_excel = leer_columna_excel(excel_path, index_col=1)
    resultados = comparar_nombres_fuzzy(nombres_pdf, nombres_excel)
    exportar_resultados_nombres(resultados)
    return len(resultados)

def procesar_comparar_documentos(carpeta: Path, excel_path: str):
    documentos_pdf = []
    for item in Path(carpeta).iterdir():
        if item.is_file() and item.suffix.lower() == ".pdf":
            texto = leer_pdf_completo(item)
            num, tipo = extraer_documento_desde_texto(texto)
            if num and tipo:
                documentos_pdf.append((num, tipo, item.name))
    documentos_excel_raw = leer_columna_excel(excel_path, index_col=0)
    documentos_excel = [normalizar_documento(d) for d in documentos_excel_raw]
    resultados = comparar_documentos_exactos(documentos_pdf, documentos_excel)
    exportar_resultados_documentos(resultados)
    return len(resultados)


def procesar_validacion_completa(carpeta: Path, excel_path: str):
    nombres_excel = leer_columna_excel(excel_path, index_col=1)
    documentos_excel_raw = leer_columna_excel(excel_path, index_col=0)
    documentos_excel = [normalizar_documento(d) for d in documentos_excel_raw]

    resultados = []
    
    # ----------------------------------------------------------------
    # 🔎 PRIMERA FASE: recorrer PDFs y comparar con Excel
    # ----------------------------------------------------------------
    for item in Path(carpeta).iterdir():
        if not (item.is_file() and item.suffix.lower() == ".pdf"):
            continue
        texto = leer_pdf_completo(item)
        nombre_extraido = extraer_nombre_desde_texto(texto) or MSG_NO_EXTRAIDO
        num_extraido, tipo_doc = extraer_documento_desde_texto(texto)
        
        # Comparar nombre
        mejor_nombre = ""
        sim_nombre = 0
        if nombre_extraido != MSG_NO_EXTRAIDO:
            nombre_norm = normalizar_texto(nombre_extraido)
            nombres_excel_norm = [normalizar_texto(n) for n in nombres_excel]
            if nombre_norm in nombres_excel_norm:
                idx = nombres_excel_norm.index(nombre_norm)
                mejor_nombre = nombres_excel[idx]
                sim_nombre = 100
            else:
                for i, ne_norm in enumerate(nombres_excel_norm):
                    s = fuzz.ratio(nombre_norm, ne_norm)
                    if s > sim_nombre:
                        sim_nombre = s
                        mejor_nombre = nombres_excel[i]
        estado_nombre, _ = evaluar_similitud(sim_nombre)

        # Comparar documento
        mejor_doc = ""
        sim_doc = 0
        if num_extraido:
            if num_extraido in documentos_excel:
                mejor_doc = num_extraido
                sim_doc = 100
            else:
                for de in documentos_excel:
                    s = fuzz.ratio(num_extraido, de)
                    if s > sim_doc:
                        sim_doc = s
                        mejor_doc = de
            # 🛑 VALIDACIÓN ESPECIAL PARA NOMBRES
            if sim_nombre < 50:
                estado_nombre = "❌ NO EXISTE EN EL REPORTE DE INSCRIPCIÓN"
                mejor_nombre = ""  # No sugerir falso match
            else:
                estado_nombre, _ = evaluar_similitud(sim_nombre)

            # 🛑 VALIDACIÓN ESPECIAL PARA DOCUMENTOS
            if sim_doc < 50:
                estado_doc = "❌ NO EXISTE EN EL REPORTE DE INSCRIPCIÓN"
                mejor_doc = ""     # No sugerir falso match
            else:
                estado_doc, _ = evaluar_similitud(sim_doc)


        resultados.append([
            item.name,
            nombre_extraido,
            mejor_nombre,
            sim_nombre,
            estado_nombre,
            num_extraido or MSG_NO_EXTRAIDO,
            tipo_doc or "NO IDENTIFICADO",
            mejor_doc,
            sim_doc,
            estado_doc
        ])

    # ----------------------------------------------------------------
    # 🟦 SEGUNDA FASE: revisar Excel persona por persona (FUERA DEL FOR)
    # ----------------------------------------------------------------
    pdf_nombres_norm = [normalizar_texto(r[1]) for r in resultados]
    pdf_docs = [r[5] for r in resultados]

    for idx, nombre in enumerate(nombres_excel):
        nombre_norm = normalizar_texto(nombre)
        documento_excel = documentos_excel[idx] if idx < len(documentos_excel) else ""

        falta_nombre = nombre_norm not in pdf_nombres_norm
        falta_doc = documento_excel not in pdf_docs if documento_excel else True

        if falta_nombre or falta_doc:
            resultados.append([
                "— SIN PDF —",
                MSG_NO_EXTRAIDO,
                nombre,
                0,
                "❌ NO TIENE CERTIFICADO",
                documento_excel if documento_excel else MSG_NO_EXTRAÍDO,
                "NO IDENTIFICADO",
                documento_excel if documento_excel else "",
                0,
                "❌ NO TIENE CERTIFICADO"
            ])

    # ----------------------------------------------------------------
    # 📤 Exportar Excel final
    # ----------------------------------------------------------------
    exportar_resultados_completos(resultados)
    return len(resultados)


# ------------------------------
# GUI (limpia pero funcional)
# ------------------------------
class ValidadorUnificado:
    def __init__(self):
        self.ventana = tk.Tk()
        self._configurar_ventana()
        self._crear_interfaz()

    def _configurar_ventana(self):
        self.ventana.title("🔍 Validador Unificado PDF vs Excel")
        self.ventana.geometry("800x700")
        self.ventana.configure(bg="#f0f0f0")
        self.ventana.resizable(True, True)
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("Title.TLabel", font=("Segoe UI", 18, "bold"), background="#f0f0f0", foreground="#2c3e50")
        style.configure("Subtitle.TLabel", font=("Segoe UI", 11), background="#f0f0f0", foreground="#34495e")
        style.configure("Action.TButton", font=("Segoe UI", 11, "bold"), padding=(15, 8))

    def _crear_interfaz(self):
        main = ttk.Frame(self.ventana, padding="30")
        main.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main, text="Validador Unificado PDF vs Excel", style="Title.TLabel").pack(pady=(0, 10))
        ttk.Label(main, text="Valida nombres y números de documento de PDFs contra reportes Excel", style="Subtitle.TLabel").pack(pady=(0, 20))

        # Nombres
        nf = ttk.LabelFrame(main, text="📝 Validación de Nombres", padding="10")
        nf.pack(fill=tk.X, pady=(0, 10))
        ttk.Button(nf, text="🔍 Extraer Nombres", command=self._thread(self._ui_extraer_nombres), style="Action.TButton").pack(side=tk.LEFT, padx=4)
        ttk.Button(nf, text="⚖️ Comparar Nombres", command=self._thread(self._ui_comparar_nombres), style="Action.TButton").pack(side=tk.LEFT, padx=4)

        # Documentos
        df = ttk.LabelFrame(main, text="🆔 Validación de Documentos", padding="10")
        df.pack(fill=tk.X, pady=(0, 10))
        ttk.Button(df, text="🔍 Extraer Documentos", command=self._thread(self._ui_extraer_documentos), style="Action.TButton").pack(side=tk.LEFT, padx=4)
        ttk.Button(df, text="⚖️ Comparar Documentos", command=self._thread(self._ui_comparar_documentos), style="Action.TButton").pack(side=tk.LEFT, padx=4)

        # Completa
        cf = ttk.LabelFrame(main, text="🎯 Validación Completa", padding="10")
        cf.pack(fill=tk.X, pady=(0, 10))
        ttk.Button(cf, text="🚀 Validación Completa", command=self._thread(self._ui_validacion_completa), style="Action.TButton").pack()

        # Progress
        self.progress_label = ttk.Label(main, text="", style="Subtitle.TLabel")
        self.progress_label.pack(pady=(10, 0))
        self.progress = ttk.Progressbar(main, mode="indeterminate")
        self.progress.pack(fill=tk.X, pady=(5, 0))

    def _thread(self, target_func):
        """Return a callable that starts target_func in a daemon thread with minimal UI changes."""
        def wrapper():
            threading.Thread(target=self._run_with_progress, args=(target_func,), daemon=True).start()
        return wrapper

    def _run_with_progress(self, func):
        try:
            self._show_progress("Procesando...")
            func()
        except Exception as e:
            logging.exception("Error en proceso de GUI: %s", e)
            messagebox.showerror("Error", f"Ocurrió un error: {e}")
        finally:
            self._hide_progress()

    def _show_progress(self, msg: str):
        self.progress_label.config(text=msg)
        self.progress.start(10)

    def _hide_progress(self):
        self.progress.stop()
        self.progress_label.config(text="")
        # allow UI to update
        self.ventana.update_idletasks()

    # UI actions
    def _ui_extraer_nombres(self):
        carpeta = filedialog.askdirectory(title="Seleccionar carpeta de PDFs", initialdir=str(Path.home()))
        if not carpeta:
            return
        ruta, cantidad = procesar_extraer_nombres(Path(carpeta))
        if ruta:
            messagebox.showinfo("Éxito", f"✅ Nombres extraídos: {cantidad}\n📁 Guardado en: {ruta}")
        else:
            messagebox.showwarning("Atención", "No se pudieron extraer nombres de ningún PDF.")

    def _ui_comparar_nombres(self):
        carpeta = filedialog.askdirectory(title="Seleccionar carpeta de PDFs", initialdir=str(Path.home()))
        if not carpeta:
            return
        archivo = filedialog.askopenfilename(title="Seleccionar archivo Excel", filetypes=[("Excel", "*.xlsx *.xls")])
        if not archivo:
            return
        procesar_comparar_nombres(Path(carpeta), archivo)

    def _ui_extraer_documentos(self):
        carpeta = filedialog.askdirectory(title="Seleccionar carpeta de PDFs", initialdir=str(Path.home()))
        if not carpeta:
            return
        ruta, cantidad = procesar_extraer_documentos(Path(carpeta))
        if ruta:
            messagebox.showinfo("Éxito", f"✅ Documentos extraídos: {cantidad}\n📁 Guardado en: {ruta}")
        else:
            messagebox.showwarning("Atención", "No se pudieron extraer documentos de ningún PDF.")

    def _ui_comparar_documentos(self):
        carpeta = filedialog.askdirectory(title="Seleccionar carpeta de PDFs", initialdir=str(Path.home()))
        if not carpeta:
            return
        archivo = filedialog.askopenfilename(title="Seleccionar archivo Excel", filetypes=[("Excel", "*.xlsx *.xls")])
        if not archivo:
            return
        procesar_comparar_documentos(Path(carpeta), archivo)

    def _ui_validacion_completa(self):
        carpeta = filedialog.askdirectory(title="Seleccionar carpeta de PDFs", initialdir=str(Path.home()))
        if not carpeta:
            return
        archivo = filedialog.askopenfilename(title="Seleccionar archivo Excel", filetypes=[("Excel", "*.xlsx *.xls")])
        if not archivo:
            return
        procesar_validacion_completa(Path(carpeta), archivo)

    def ejecutar(self):
        CARPETA_DESCARGAS.mkdir(parents=True, exist_ok=True)
        self.ventana.mainloop()

# ------------------------------
# Punto de entrada
# ------------------------------
if __name__ == "__main__":
    app = ValidadorUnificado()
    app.ejecutar()
