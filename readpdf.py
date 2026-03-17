import os
import shutil
import re
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook  # 📘 Import necesario para leer Excel

# ------------------------------------------------------------
# 🧩 FUNCIÓN 1: Mover PDFs a sus carpetas por número de ficha
# ------------------------------------------------------------
def mover_pdfs():
    carpeta_pdfs = filedialog.askdirectory(title="Selecciona la carpeta donde están los PDFs")
    if not carpeta_pdfs:
        return

    carpeta_destino = filedialog.askdirectory(title="Selecciona la carpeta donde están las fichas")
    if not carpeta_destino:
        return

    movidos = 0
    no_encontrados = 0
    sin_ficha = []

    patron_ficha = re.compile(r'(\d{7})')  # Buscar 7 dígitos consecutivos

    for archivo in os.listdir(carpeta_pdfs):
        if archivo.lower().endswith(".pdf"):
            ruta_pdf = os.path.join(carpeta_pdfs, archivo)
            nombre_limpio = archivo.replace(" ", "").replace("_", "").replace("-", "")
            coincidencia = patron_ficha.search(nombre_limpio)

            if coincidencia:
                ficha = coincidencia.group(1)
                carpeta_ficha = os.path.join(carpeta_destino, ficha)

                if os.path.exists(carpeta_ficha):
                    destino = os.path.join(carpeta_ficha, archivo)
                    try:
                        shutil.move(ruta_pdf, destino)
                        movidos += 1
                    except Exception as e:
                        sin_ficha.append(f"{archivo} (error: {e})")
                        no_encontrados += 1
                else:
                    no_encontrados += 1
                    sin_ficha.append(f"{archivo} (ficha {ficha} sin carpeta)")
            else:
                no_encontrados += 1
                sin_ficha.append(f"{archivo} (sin número de ficha detectado)")

    mensaje = f"✅ Archivos movidos: {movidos}\n❌ No movidos: {no_encontrados}"
    if sin_ficha:
        mensaje += "\n\n⚠️ Archivos no procesados:\n" + "\n".join(sin_ficha[:15])

    messagebox.showinfo("Resultado del proceso", mensaje)


# ------------------------------------------------------------
# 🧩 FUNCIÓN 2: Copiar un Excel a TODAS las carpetas con 7 números
# ------------------------------------------------------------
def copiar_excel_a_fichas():
    archivo_excel = filedialog.askopenfilename(
        title="Selecciona el archivo Excel a copiar",
        filetypes=[("Archivos Excel", "*.xlsx *.xls")]
    )
    if not archivo_excel:
        return

    carpeta_base = filedialog.askdirectory(title="Selecciona la carpeta donde están las fichas")
    if not carpeta_base:
        return

    patron_ficha = re.compile(r'^\d{7}$')
    carpetas_fichas = [
        os.path.join(carpeta_base, c)
        for c in os.listdir(carpeta_base)
        if os.path.isdir(os.path.join(carpeta_base, c)) and patron_ficha.match(c)
    ]

    if not carpetas_fichas:
        messagebox.showwarning("Sin carpetas válidas", "No se encontraron carpetas con 7 números en el nombre.")
        return

    copiados = 0
    errores = []

    for carpeta in carpetas_fichas:
        try:
            nombre_archivo = os.path.basename(archivo_excel)
            destino = os.path.join(carpeta, nombre_archivo)
            shutil.copy(archivo_excel, destino)
            copiados += 1
        except Exception as e:
            errores.append(f"{carpeta}: {e}")

    mensaje = f"✅ Excel copiado en {copiados} carpetas con número de ficha."
    if errores:
        mensaje += f"\n❌ Errores en {len(errores)} carpetas:\n" + "\n".join(errores[:10])

    messagebox.showinfo("Resultado del proceso", mensaje)


# ------------------------------------------------------------
# 🧩 FUNCIÓN 3 (versión segura): Renombrar Excels sin alterar formato
# ------------------------------------------------------------
def renombrar_excels_por_ficha():
    from openpyxl import load_workbook
    import xlrd  # 👈 se usa solo para leer .xls sin modificar

    carpeta_excels = filedialog.askdirectory(title="Selecciona la carpeta donde están los reportes de inscripción")
    if not carpeta_excels:
        return

    renombrados = 0
    errores = []

    patron_reporte = re.compile(r"reporte[\s_]*inscripcion", re.IGNORECASE)

    for archivo in os.listdir(carpeta_excels):
        ruta_archivo = os.path.join(carpeta_excels, archivo)
        nombre_lower = archivo.lower()

        if patron_reporte.search(nombre_lower) and (nombre_lower.endswith(".xlsx") or nombre_lower.endswith(".xls")):
            try:
                ficha = None

                # Leer celda B3 según el tipo de archivo
                if nombre_lower.endswith(".xlsx"):
                    wb = load_workbook(ruta_archivo, data_only=True)
                    hoja = wb.active
                    ficha = hoja["B3"].value
                    wb.close()
                elif nombre_lower.endswith(".xls"):
                    try:
                        libro = xlrd.open_workbook(ruta_archivo)
                        hoja = libro.sheet_by_index(0)
                        ficha = hoja.cell_value(2, 1)  # fila 2, columna 1 = B3
                    except Exception as e:
                        errores.append(f"{archivo}: error al leer .xls ({e})")
                        continue

                # Validar la ficha
                if ficha and str(ficha).isdigit() and len(str(ficha)) == 7:
                    nuevo_nombre = f"reporte_inscripcion {int(ficha)}{os.path.splitext(archivo)[1]}"
                    nueva_ruta = os.path.join(carpeta_excels, nuevo_nombre)

                    if not os.path.exists(nueva_ruta):
                        os.rename(ruta_archivo, nueva_ruta)
                        renombrados += 1
                    else:
                        errores.append(f"{archivo} → {nuevo_nombre} (ya existe)")
                else:
                    errores.append(f"{archivo}: ficha inválida en B3 ({ficha})")

            except Exception as e:
                errores.append(f"{archivo}: {e}")

    mensaje = f"✅ Archivos renombrados: {renombrados}"
    if errores:
        mensaje += f"\n❌ Errores en {len(errores)} archivos:\n" + "\n".join(errores[:10])

    messagebox.showinfo("Resultado del proceso", mensaje)

    # ------------------------------------------------------------
# 🧩 FUNCIÓN 4: Mover reportes_inscripcion.xlsx a su carpeta según número de ficha
# ------------------------------------------------------------
def mover_reportes_inscripcion():
    carpeta_excels = filedialog.askdirectory(title="Selecciona la carpeta donde están los reportes de inscripción")
    if not carpeta_excels:
        return

    carpeta_destino = filedialog.askdirectory(title="Selecciona la carpeta donde están las fichas")
    if not carpeta_destino:
        return

    patron_ficha = re.compile(r"(\d{7})")  # número de ficha (7 dígitos)
    patron_reporte = re.compile(r"reporte[\s_]*inscripcion", re.IGNORECASE)

    movidos = 0
    reemplazados = 0
    no_encontrados = 0
    sin_ficha = []

    for archivo in os.listdir(carpeta_excels):
        ruta_excel = os.path.join(carpeta_excels, archivo)
        nombre_lower = archivo.lower()

        # Solo procesar archivos que contengan "reporte inscripcion" en el nombre
        if patron_reporte.search(nombre_lower) and (nombre_lower.endswith(".xlsx") or nombre_lower.endswith(".xls")):
            nombre_limpio = archivo.replace(" ", "").replace("_", "").replace("-", "")
            coincidencia = patron_ficha.search(nombre_limpio)

            if coincidencia:
                ficha = coincidencia.group(1)
                carpeta_ficha = os.path.join(carpeta_destino, ficha)

                if os.path.exists(carpeta_ficha):
                    destino = os.path.join(carpeta_ficha, archivo)
                    try:
                        # ⚠️ Si ya existe un archivo con el mismo nombre, eliminarlo
                        if os.path.exists(destino):
                            os.remove(destino)
                            reemplazados += 1

                        shutil.move(ruta_excel, destino)
                        movidos += 1

                    except Exception as e:
                        no_encontrados += 1
                        sin_ficha.append(f"{archivo} (error al mover: {e})")
                else:
                    no_encontrados += 1
                    sin_ficha.append(f"{archivo} (ficha {ficha} sin carpeta)")
            else:
                no_encontrados += 1
                sin_ficha.append(f"{archivo} (sin número de ficha detectado)")

    # 📊 Resumen del proceso
    mensaje = (
        f"✅ Reportes movidos correctamente: {movidos}\n"
        f"🔁 Archivos reemplazados: {reemplazados}\n"
        f"❌ No movidos: {no_encontrados}"
    )
    if sin_ficha:
        mensaje += "\n\n⚠️ Archivos no procesados:\n" + "\n".join(sin_ficha[:15])

    messagebox.showinfo("Resultado del proceso", mensaje)

# ------------------------------------------------------------
# 🖥️ INTERFAZ GRÁFICA PRINCIPAL
# ------------------------------------------------------------
root = tk.Tk()
root.title("Organizador de Fichas PDF y Excel")
root.geometry("500x400")

tk.Label(root, text="📂 Organizador de Fichas", font=("Arial", 14, "bold")).pack(pady=10)
tk.Label(
    root,
    text="1️⃣ Mover PDFs según número de ficha\n2️⃣ Copiar Excel en carpetas de fichas (7 números)\n3️⃣ Renombrar Excels con la ficha (celda B3)",
    justify="center",
    font=("Arial", 10)
).pack(pady=10)

tk.Button(
    root, text="📄 Mover PDFs a carpetas de fichas", command=mover_pdfs,
    bg="#4CAF50", fg="white", font=("Arial", 11), width=40, height=2
).pack(pady=5)

tk.Button(
    root, text="📊 Copiar Excel en carpetas de fichas (7 dígitos)", command=copiar_excel_a_fichas,
    bg="#2196F3", fg="white", font=("Arial", 11), width=40, height=2
).pack(pady=5)

tk.Button(
    root, text="🧾 Renombrar Excels según ficha (celda B3)", command=renombrar_excels_por_ficha,
    bg="#FF9800", fg="white", font=("Arial", 11), width=40, height=2
).pack(pady=5)
tk.Button(
    root, text="📁 Mover reportes_inscripcion a carpetas de fichas", command=mover_reportes_inscripcion,
    bg="#9C27B0", fg="white", font=("Arial", 11), width=40, height=2
).pack(pady=5)

tk.Label(root, text="💡 Desarrollado para automatizar el proceso de fichas", font=("Arial", 8), fg="gray").pack(side="bottom", pady=10)

root.mainloop()