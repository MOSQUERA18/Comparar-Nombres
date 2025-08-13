import re
import unicodedata
import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
from typing import List, Tuple, Optional, Any

import pandas as pd
import pdfplumber
from fuzzywuzzy import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# --- Constantes ---
CARPETA_DESCARGAS = Path.home() / "Downloads"
NOMBRE_ARCHIVO_SALIDA_COMPARACION = "comparacion_documentos.xlsx"
NOMBRE_ARCHIVO_SALIDA_EXTRACCION = "documentos_extraidos.xlsx"

PATRON_CEDULA_ADULTO = re.compile(r"Cédula de Ciudadanía:\s*([0-9]{1,3}(?:\.[0-9]{3})*\.[0-9]{3})", re.IGNORECASE)
PATRON_NUIP_MENOR = re.compile(r"Número Único de Identificación Personal\s+([0-9]+)", re.IGNORECASE)
PATRON_RUMV_PPT = re.compile(r"número de RUMV\s+([0-9]+)", re.IGNORECASE)

# Estilos para Excel
VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
AMARILLO = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ROJO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# --- Funciones de Utilidad ---

def normalizar_documento(documento: str) -> str:
    """
    Normaliza un número de documento: elimina puntos, comas, espacios y caracteres especiales.
    """
    if not isinstance(documento, str):
        documento = str(documento)
        
    documento = re.sub(r"[^0-9]", "", documento)
    return documento.strip()

# --- Lógica Principal de Procesamiento ---

def extraer_documento_desde_pdf(pdf_path: Path) -> Tuple[Optional[str], Optional[str]]:
    """
    Extrae un número de documento de un archivo PDF utilizando patrones regex.
    Retorna (numero_documento, tipo_documento) o (None, None) si no encuentra nada.
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            texto_completo = ""
            for pagina in pdf.pages:
                if pagina.extract_text():
                    texto_completo += pagina.extract_text() + "\n"

            match_cedula = PATRON_CEDULA_ADULTO.search(texto_completo)
            if match_cedula:
                numero_normalizado = normalizar_documento(match_cedula.group(1))
                return numero_normalizado, "CEDULA_ADULTO"

            match_nuip = PATRON_NUIP_MENOR.search(texto_completo)
            if match_nuip:
                numero_normalizado = normalizar_documento(match_nuip.group(1))
                return numero_normalizado, "NUIP_MENOR"

            match_rumv = PATRON_RUMV_PPT.search(texto_completo)
            if match_rumv:
                numero_normalizado = normalizar_documento(match_rumv.group(1))
                return numero_normalizado, "RUMV_PPT"
                
    except Exception as e:
        print(f"Error al procesar el PDF {pdf_path.name}: {e}")
    return None, None

def leer_documentos_desde_excel(ruta_excel: Path) -> List[str]:
    """
    Lee números de documento de la columna A en un archivo Excel.
    Asume que los documentos están en la columna A (índice 0) y omite las primeras 6 filas.
    """
    try:
        df = pd.read_excel(ruta_excel, usecols=[0], skiprows=6, header=None)
        documentos_raw = df.iloc[:, 0].dropna().astype(str).tolist()
        return [normalizar_documento(doc) for doc in documentos_raw]
    except Exception as e:
        messagebox.showerror("Error de lectura Excel", f"No se pudo leer el archivo Excel:\n{e}")
        return []

def _procesar_carpeta_pdfs(carpeta_path: Path) -> Tuple[List[Tuple[str, str, str]], List[str]]:
    """
    Procesa todos los archivos PDF en una carpeta, extrayendo números de documento.
    Helper function.

    Returns:
        Una tupla (documentos_extraidos, archivos_no_leidos).
        documentos_extraidos es una lista de tuplas (numero_documento, tipo_documento, nombre_archivo).
    """
    documentos_extraidos: List[Tuple[str, str, str]] = []
    archivos_no_leidos: List[str] = []

    for item in carpeta_path.iterdir():
        if item.is_file() and item.suffix.lower() == ".pdf":
            numero_doc, tipo_doc = extraer_documento_desde_pdf(item)
            if numero_doc and tipo_doc:
                documentos_extraidos.append((numero_doc, tipo_doc, item.name))
            else:
                archivos_no_leidos.append(item.name)
    return documentos_extraidos, archivos_no_leidos

def encontrar_mejor_coincidencia(documento_pdf: str, documentos_excel: List[str]) -> Tuple[str, int, str]:
    """
    Encuentra la mejor coincidencia para un documento PDF en la lista de Excel.
    Retorna (mejor_match, similitud, estado).
    """
    if documento_pdf in documentos_excel:
        return documento_pdf, 100, "✔ EXACTA"
    
    mejor_similitud = 0
    mejor_match = ""
    
    for doc_excel in documentos_excel:
        similitud = fuzz.ratio(documento_pdf, doc_excel)
        if similitud > mejor_similitud:
            mejor_similitud = similitud
            mejor_match = doc_excel
    
    if mejor_similitud >= 90:
        estado = "⚠️ ALTA SIMILITUD"
    elif mejor_similitud >= 70:
        estado = "⚠️ MEDIA SIMILITUD"
    elif mejor_similitud >= 50:
        estado = "⚠️ BAJA SIMILITUD"
    else:
        estado = "❌ SIN COINCIDENCIA"
    
    return mejor_match, mejor_similitud, estado

def comparar_documentos_pdf_excel(documentos_pdf: List[Tuple[str, str, str]], documentos_excel: List[str]) -> List[List[Any]]:
    """
    Compara una lista de documentos extraídos de PDFs con una lista de documentos de un Excel.
    Utiliza coincidencia exacta y fuzzy matching para números de documento.
    """
    resultados: List[List[Any]] = []
    documentos_excel_disponibles = documentos_excel.copy()

    for numero_doc_pdf, tipo_doc, nombre_archivo in documentos_pdf:
        if not numero_doc_pdf:
            resultados.append([nombre_archivo, numero_doc_pdf, tipo_doc, "DOCUMENTO VACIO", 0, "❌ DOCUMENTO VACIO"])
            continue

        mejor_match, similitud, estado = encontrar_mejor_coincidencia(numero_doc_pdf, documentos_excel_disponibles)
        
        # Si es coincidencia exacta, remover de la lista disponible
        if similitud == 100:
            documentos_excel_disponibles.remove(numero_doc_pdf)

        resultados.append([nombre_archivo, numero_doc_pdf, tipo_doc, mejor_match, similitud, estado])
        
    return resultados

def exportar_resultados_a_excel_con_colores(resultados: List[List[Any]], ruta_salida_completa: Path) -> None:
    """
    Exporta los resultados de la comparación a un archivo Excel, coloreando las filas.
    """
    df_resultados = pd.DataFrame(resultados, columns=[
        "Archivo PDF", "Documento Extraído", "Tipo Documento", 
        "Mejor Coincidencia Excel", "% Similitud", "Estado"
    ])
    
    try:
        df_resultados.to_excel(ruta_salida_completa, index=False)
        
        wb = load_workbook(ruta_salida_completa)
        ws = wb.active
        
        for row_idx in range(2, ws.max_row + 1):
            estado_celda = ws[f"F{row_idx}"].value 
            
            if "EXACTA" in str(estado_celda):
                fill_color = VERDE
            elif "SIMILITUD" in str(estado_celda):
                fill_color = AMARILLO
            else:
                fill_color = ROJO

            if fill_color:
                for col_letter in ["A", "B", "C", "D", "E", "F"]:
                    ws[f"{col_letter}{row_idx}"].fill = fill_color
            
        wb.save(ruta_salida_completa)
        messagebox.showinfo("Exportación Exitosa", f"Comparación de documentos exportada y formateada en:\n{ruta_salida_completa}")
    except Exception as e:
        messagebox.showerror("Error de Exportación", f"No se pudo exportar el archivo Excel:\n{e}")


# --- Funciones de Interfaz de Usuario (Callbacks) ---

def _accion_extraer_documentos() -> None:
    """
    Callback para el botón "Extraer documentos de PDFs".
    Pide al usuario una carpeta, procesa los PDFs y guarda los documentos en un Excel.
    """
    carpeta_seleccionada = filedialog.askdirectory(title="Seleccionar carpeta de PDFs", initialdir=str(Path.home()))
    if not carpeta_seleccionada:
        return

    carpeta_pdfs_path = Path(carpeta_seleccionada)
    documentos_extraidos, no_leidos = _procesar_carpeta_pdfs(carpeta_pdfs_path)

    if not documentos_extraidos and not no_leidos:
        messagebox.showinfo("Resultado", "No se encontraron archivos PDF en la carpeta seleccionada.")
        return
    
    if documentos_extraidos:
        datos_para_excel = []
        for numero_doc, tipo_doc, nombre_archivo in documentos_extraidos:
            datos_para_excel.append([nombre_archivo, numero_doc, tipo_doc])
        
        df_documentos = pd.DataFrame(datos_para_excel, columns=["Archivo PDF", "Número Documento", "Tipo Documento"])
        
        ruta_salida_extraccion = CARPETA_DESCARGAS / NOMBRE_ARCHIVO_SALIDA_EXTRACCION
        try:
            df_documentos.to_excel(ruta_salida_extraccion, index=False)
            msg = f"Documentos extraídos y guardados en:\n{ruta_salida_extraccion}"
        except Exception as e:
            msg = f"Error al guardar documentos extraídos:\n{e}"
            messagebox.showerror("Error al Guardar", msg)
            return
    else:
        msg = "No se pudieron extraer documentos de ningún PDF."

    if no_leidos:
        msg += f"\n\nArchivos PDF no procesados o sin documento encontrado:\n- " + "\n- ".join(no_leidos)
    
    messagebox.showinfo("Resultado de Extracción", msg)

def _accion_comparar_documentos() -> None:
    """
    Callback para el botón "Comparar con Excel".
    Pide carpeta de PDFs y archivo Excel, compara documentos y exporta resultados.
    """
    carpeta_seleccionada = filedialog.askdirectory(title="Seleccionar carpeta de PDFs", initialdir=str(Path.home()))
    if not carpeta_seleccionada:
        return
    
    archivo_excel_seleccionado = filedialog.askopenfilename(
        title="Seleccionar archivo Excel de validación",
        filetypes=[("Archivos Excel", "*.xlsx *.xls")],
        initialdir=str(Path.home())
    )
    if not archivo_excel_seleccionado:
        return

    carpeta_pdfs_path = Path(carpeta_seleccionada)
    excel_validacion_path = Path(archivo_excel_seleccionado)

    documentos_pdf, no_leidos_pdf = _procesar_carpeta_pdfs(carpeta_pdfs_path)
    
    if not documentos_pdf:
        messagebox.showwarning("Atención", "No se extrajeron documentos de ningún PDF en la carpeta seleccionada.")
        if no_leidos_pdf:
             messagebox.showinfo("PDFs no leídos", "Archivos PDF no procesados:\n\n" + "\n".join(no_leidos_pdf))
        return

    documentos_excel = leer_documentos_desde_excel(excel_validacion_path)
    if not documentos_excel:
        messagebox.showwarning("Atención", "No se encontraron documentos en el archivo Excel o hubo un error al leerlo.")
        return

    resultados_comparacion = comparar_documentos_pdf_excel(documentos_pdf, documentos_excel)
    
    ruta_salida_comparacion = CARPETA_DESCARGAS / NOMBRE_ARCHIVO_SALIDA_COMPARACION
    exportar_resultados_a_excel_con_colores(resultados_comparacion, ruta_salida_comparacion)

    if no_leidos_pdf:
        messagebox.showwarning("Atención - PDFs no leídos", 
                               "No se pudo extraer documento de los siguientes archivos PDF:\n\n" + "\n".join(no_leidos_pdf))

# --- Configuración y Lanzamiento de la GUI ---

def lanzar_gui() -> None:
    """
    Crea y lanza la interfaz gráfica de usuario (GUI) con Tkinter.
    """
    ventana = tk.Tk()
    ventana.title("Validador de Números de Documento PDF vs Excel")
    ventana.geometry("500x320")
    
    # Estilo simple para los botones
    estilo_boton = {'relief': tk.RAISED, 'borderwidth': 2, 'font': ('Arial', 10, 'bold'), 'height': 2, 'width': 40, 'pady': 5}

    label_titulo = tk.Label(ventana, text="Herramienta de Validación de Documentos", font=('Arial', 14, 'bold'))
    label_titulo.pack(pady=(10,15))

    label_desc = tk.Label(
        ventana,
        text="Valida números de documento de:\n• Cédulas de Ciudadanía (adultos)\n• NUIP (menores de edad)\n• RUMV (migrantes venezolanos)",
        font=('Arial', 9),
        justify=tk.CENTER
    )
    label_desc.pack(pady=(0,15))

    boton_extraer = tk.Button(ventana, text="Extraer Documentos de PDFs a Excel", command=_accion_extraer_documentos, **estilo_boton)
    boton_extraer.pack(pady=10)

    boton_comparar = tk.Button(ventana, text="Comparar Documentos de PDFs con Excel", command=_accion_comparar_documentos, **estilo_boton)
    boton_comparar.pack(pady=10)

    label_colores = tk.Label(
        ventana,
        text="Colores del reporte:\n🟢 Verde: Coincidencia exacta | 🟡 Amarillo: Similitud parcial | 🔴 Rojo: Sin coincidencia",
        font=('Arial', 8),
        fg='gray',
        justify=tk.CENTER
    )
    label_colores.pack(pady=(10,5))

    label_info = tk.Label(
        ventana,
        text="Los documentos en Excel deben estar en columna A desde fila 7",
        font=('Arial', 8),
        fg='gray'
    )
    label_info.pack(pady=(5,0))

    ventana.mainloop()

# --- Punto de Entrada del Script ---

if __name__ == "__main__":
    # Asegurarse que la carpeta de descargas exista, si no, crearla.
    CARPETA_DESCARGAS.mkdir(parents=True, exist_ok=True) 
    lanzar_gui()
