import re
import unicodedata
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from typing import List, Tuple, Optional, Any
import threading

import pandas as pd
import pdfplumber
from fuzzywuzzy import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# --- Constantes ---
CARPETA_DESCARGAS = Path.home() / "Downloads"

# Patrones Regex para nombres
PATRON_GENERAL_NOMBRE = re.compile(r"A nombre de:\s*(.+?)\s*Estado:", re.IGNORECASE)
PATRON_REGISTRO_CIVIL = re.compile(r"Registro Civil,\s*(.*?)\s*tiene inscrito", re.IGNORECASE | re.DOTALL)
PATRON_MIGRACION_NOMBRE = re.compile(r"el migrante venezolano\s+([A-ZÁÉÍÓÚÑ ]{5,})\s+surtió", re.IGNORECASE | re.DOTALL)

# Patrones Regex para documentos
PATRON_CEDULA_ADULTO = re.compile(r"Cédula de Ciudadanía:\s*([0-9]{1,3}(?:\.[0-9]{3})*\.[0-9]{3})", re.IGNORECASE)
PATRON_NUIP_MENOR = re.compile(r"Número Único de Identificación Personal\s+([0-9]+)", re.IGNORECASE)
PATRON_RUMV_PPT = re.compile(r"número de RUMV\s+([0-9]+)", re.IGNORECASE)

# Estilos para Excel
VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
AMARILLO = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ROJO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
AZUL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

# --- Funciones de Utilidad ---

def normalizar_texto(texto: str) -> str:
    """Normaliza un texto: convierte a mayúsculas, elimina tildes, puntuación y espacios extra."""
    if not isinstance(texto, str):
        texto = str(texto)
    texto = texto.strip().upper()
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('utf-8')
    texto = re.sub(r"[^A-Z\s]", "", texto)
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()

def normalizar_documento(documento: str) -> str:
    """Normaliza un número de documento: elimina puntos, comas, espacios y caracteres especiales."""
    if not isinstance(documento, str):
        documento = str(documento)
    documento = re.sub(r"[^0-9]", "", documento)
    return documento.strip()

def invertir_nombre(nombre: str) -> str:
    """Invierte un nombre si consta de múltiples partes."""
    partes = nombre.strip().split()
    if len(partes) >= 2:
        mitad = len(partes) // 2
        return " ".join(partes[mitad:] + partes[:mitad])
    return nombre

# --- Lógica de Extracción ---

def extraer_nombre_desde_pdf(pdf_path: Path) -> Optional[str]:
    """Extrae un nombre de un archivo PDF utilizando patrones regex."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            texto_completo = ""
            for pagina in pdf.pages:
                if pagina.extract_text():
                    texto_completo += pagina.extract_text() + "\n"

            match_general = PATRON_GENERAL_NOMBRE.search(texto_completo)
            if match_general:
                return match_general.group(1).strip()

            match_registro_civil = PATRON_REGISTRO_CIVIL.search(texto_completo)
            if match_registro_civil:
                return invertir_nombre(match_registro_civil.group(1).strip())

            match_migracion = PATRON_MIGRACION_NOMBRE.search(texto_completo)
            if match_migracion:
                return match_migracion.group(1).strip()
                
    except Exception as e:
        print(f"Error al procesar el PDF {pdf_path.name}: {e}")
    return None

def extraer_documento_desde_pdf(pdf_path: Path) -> Tuple[Optional[str], Optional[str]]:
    """Extrae un número de documento de un archivo PDF."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            texto_completo = ""
            for pagina in pdf.pages:
                if pagina.extract_text():
                    texto_completo += pagina.extract_text() + "\n"

            match_cedula = PATRON_CEDULA_ADULTO.search(texto_completo)
            if match_cedula:
                numero_normalizado = normalizar_documento(match_cedula.group(1))
                return numero_normalizado, "CEDULA_ADULTO"

            match_nuip = PATRON_NUIP_MENOR.search(texto_completo)
            if match_nuip:
                numero_normalizado = normalizar_documento(match_nuip.group(1))
                return numero_normalizado, "NUIP_MENOR"

            match_rumv = PATRON_RUMV_PPT.search(texto_completo)
            if match_rumv:
                numero_normalizado = normalizar_documento(match_rumv.group(1))
                return numero_normalizado, "RUMV_PPT"
                
    except Exception as e:
        print(f"Error al procesar el PDF {pdf_path.name}: {e}")
    return None, None

# --- Lógica de Comparación ---

def comparar_nombres_fuzzy(nombres_pdf: List[str], nombres_excel: List[str]) -> List[List[Any]]:
    """Compara nombres usando fuzzy matching."""
    resultados = []
    nombres_excel_normalizados = [normalizar_texto(n) for n in nombres_excel]

    for nombre_pdf_original in nombres_pdf:
        nombre_pdf_normalizado = normalizar_texto(nombre_pdf_original)
        
        if not nombre_pdf_normalizado:
            resultados.append([nombre_pdf_original, "NOMBRE PDF VACIO/INVALIDO", 0, "❌"])
            continue

        mejor_match_excel_original = ""
        mejor_score = 0
        
        try:
            idx = nombres_excel_normalizados.index(nombre_pdf_normalizado)
            mejor_match_excel_original = nombres_excel[idx]
            mejor_score = 100
        except ValueError:
            for i, nombre_excel_norm in enumerate(nombres_excel_normalizados):
                if not nombre_excel_norm:
                    continue
                score = fuzz.ratio(nombre_pdf_normalizado, nombre_excel_norm)
                if score > mejor_score:
                    mejor_score = score
                    mejor_match_excel_original = nombres_excel[i]
        
        if mejor_score == 100:
            estado = "✔ EXACTA"
        elif mejor_score >= 90:
            estado = f"⚠️ ALTA ({mejor_score}%)"
        elif mejor_score >= 70:
            estado = f"⚠️ MEDIA ({mejor_score}%)"
        elif mejor_score >= 50:
            estado = f"⚠️ BAJA ({mejor_score}%)"
        else:
            estado = "❌ SIN COINCIDENCIA"

        resultados.append([nombre_pdf_original, mejor_match_excel_original, mejor_score, estado])
        
    return resultados

def comparar_documentos_exactos(documentos_pdf: List[Tuple[str, str, str]], documentos_excel: List[str]) -> List[List[Any]]:
    """Compara documentos usando coincidencia exacta y fuzzy matching."""
    resultados = []
    documentos_excel_disponibles = documentos_excel.copy()

    for numero_doc_pdf, tipo_doc, nombre_archivo in documentos_pdf:
        if not numero_doc_pdf:
            resultados.append([nombre_archivo, numero_doc_pdf, tipo_doc, "DOCUMENTO VACIO", 0, "❌ DOCUMENTO VACIO"])
            continue

        if numero_doc_pdf in documentos_excel_disponibles:
            documentos_excel_disponibles.remove(numero_doc_pdf)
            resultados.append([nombre_archivo, numero_doc_pdf, tipo_doc, numero_doc_pdf, 100, "✔ EXACTA"])
            continue
        
        mejor_similitud = 0
        mejor_match = ""
        
        for doc_excel in documentos_excel:
            similitud = fuzz.ratio(numero_doc_pdf, doc_excel)
            if similitud > mejor_similitud:
                mejor_similitud = similitud
                mejor_match = doc_excel
        
        if mejor_similitud >= 90:
            estado = f"⚠️ ALTA ({mejor_similitud}%)"
        elif mejor_similitud >= 70:
            estado = f"⚠️ MEDIA ({mejor_similitud}%)"
        elif mejor_similitud >= 50:
            estado = f"⚠️ BAJA ({mejor_similitud}%)"
        else:
            estado = "❌ SIN COINCIDENCIA"

        resultados.append([nombre_archivo, numero_doc_pdf, tipo_doc, mejor_match, mejor_similitud, estado])
        
    return resultados

# --- Clase Principal de la Aplicación ---

class ValidadorUnificado:
    def __init__(self):
        self.ventana = tk.Tk()
        self.configurar_ventana()
        self.crear_interfaz()
        
    def configurar_ventana(self):
        """Configura la ventana principal con estilo moderno."""
        self.ventana.title("🔍 Validador Unificado PDF vs Excel")
        self.ventana.geometry("800x700")
        self.ventana.configure(bg='#f0f0f0')
        self.ventana.resizable(True, True)
        
        # Configurar estilo ttk
        style = ttk.Style()
        style.theme_use('clam')
        
        # Personalizar colores
        style.configure('Title.TLabel', font=('Segoe UI', 18, 'bold'), background='#f0f0f0', foreground='#2c3e50')
        style.configure('Subtitle.TLabel', font=('Segoe UI', 11), background='#f0f0f0', foreground='#34495e')
        style.configure('Modern.TButton', font=('Segoe UI', 10, 'bold'), padding=(20, 10))
        style.configure('Action.TButton', font=('Segoe UI', 11, 'bold'), padding=(15, 8))
        
    def crear_interfaz(self):
        """Crea la interfaz gráfica moderna."""
        # Frame principal con padding
        main_frame = ttk.Frame(self.ventana, padding="30")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Título principal
        title_label = ttk.Label(main_frame, text="Validador Unificado PDF vs Excel", style='Title.TLabel')
        title_label.pack(pady=(0, 10))
        
        subtitle_label = ttk.Label(main_frame, text="Valida nombres y números de documento de PDFs contra reportes Excel", style='Subtitle.TLabel')
        subtitle_label.pack(pady=(0, 30))
        
        # Frame para botones principales
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.pack(fill=tk.X, pady=(0, 30))
        
        # Sección de Nombres
        nombres_frame = ttk.LabelFrame(buttons_frame, text="📝 Validación de Nombres", padding="20")
        nombres_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(nombres_frame, text="Extrae y compara nombres de documentos PDF usando fuzzy matching", style='Subtitle.TLabel').pack(pady=(0, 10))
        
        nombres_buttons_frame = ttk.Frame(nombres_frame)
        nombres_buttons_frame.pack(fill=tk.X)
        
        ttk.Button(nombres_buttons_frame, text="🔍 Extraer Nombres", command=self.extraer_nombres, style='Action.TButton').pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(nombres_buttons_frame, text="⚖️ Comparar Nombres", command=self.comparar_nombres, style='Action.TButton').pack(side=tk.LEFT)
        
        # Sección de Documentos
        documentos_frame = ttk.LabelFrame(buttons_frame, text="🆔 Validación de Documentos", padding="20")
        documentos_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(documentos_frame, text="Extrae y compara números de documento (Cédulas, NUIP, RUMV)", style='Subtitle.TLabel').pack(pady=(0, 10))
        
        documentos_buttons_frame = ttk.Frame(documentos_frame)
        documentos_buttons_frame.pack(fill=tk.X)
        
        ttk.Button(documentos_buttons_frame, text="🔍 Extraer Documentos", command=self.extraer_documentos, style='Action.TButton').pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(documentos_buttons_frame, text="⚖️ Comparar Documentos", command=self.comparar_documentos, style='Action.TButton').pack(side=tk.LEFT)
        
        # Sección de Validación Completa
        completa_frame = ttk.LabelFrame(buttons_frame, text="🎯 Validación Completa", padding="20")
        completa_frame.pack(fill=tk.X, pady=(0, 20))
        
        ttk.Label(completa_frame, text="Valida tanto nombres como documentos en un solo proceso", style='Subtitle.TLabel').pack(pady=(0, 10))
        
        ttk.Button(completa_frame, text="🚀 Validación Completa", command=self.validacion_completa, style='Modern.TButton').pack()
        
        # Área de progreso
        self.progress_frame = ttk.Frame(main_frame)
        self.progress_frame.pack(fill=tk.X, pady=(0, 20))
        
        self.progress_label = ttk.Label(self.progress_frame, text="", style='Subtitle.TLabel')
        self.progress_label.pack()
        
        self.progress_bar = ttk.Progressbar(self.progress_frame, mode='indeterminate')
        self.progress_bar.pack(fill=tk.X, pady=(5, 0))
        
        # Información adicional
        info_frame = ttk.LabelFrame(main_frame, text="ℹ️ Información", padding="15")
        info_frame.pack(fill=tk.X)
        
        info_text = """📋 Tipos de documento soportados:
• Cédulas de Ciudadanía (adultos)
• NUIP - Número Único de Identificación Personal (menores)
• RUMV - Registro Único de Migrantes Venezolanos (PPT)

📊 Colores del reporte Excel:
🟢 Verde: Coincidencia exacta (100%)
🟡 Amarillo: Similitud parcial (50-99%)
🔴 Rojo: Sin coincidencia (<50%)

📁 Los archivos se guardan en la carpeta Descargas"""
        
        ttk.Label(info_frame, text=info_text, style='Subtitle.TLabel', justify=tk.LEFT).pack(anchor=tk.W)
    
    def mostrar_progreso(self, mensaje: str):
        """Muestra barra de progreso con mensaje."""
        self.progress_label.config(text=mensaje)
        self.progress_bar.start(10)
        self.ventana.update()
    
    def ocultar_progreso(self):
        """Oculta la barra de progreso."""
        self.progress_bar.stop()
        self.progress_label.config(text="")
        self.ventana.update()
    
    def extraer_nombres(self):
        """Extrae nombres de PDFs y los guarda en Excel."""
        carpeta_seleccionada = filedialog.askdirectory(title="Seleccionar carpeta de PDFs", initialdir=str(Path.home()))
        if not carpeta_seleccionada:
            return
        
        def proceso():
            self.mostrar_progreso("Extrayendo nombres de PDFs...")
            
            carpeta_pdfs_path = Path(carpeta_seleccionada)
            nombres_extraidos = []
            archivos_no_leidos = []
            
            for item in carpeta_pdfs_path.iterdir():
                if item.is_file() and item.suffix.lower() == ".pdf":
                    nombre = extraer_nombre_desde_pdf(item)
                    if nombre:
                        nombres_extraidos.append([item.name, nombre])
                    else:
                        archivos_no_leidos.append(item.name)
            
            if nombres_extraidos:
                df_nombres = pd.DataFrame(nombres_extraidos, columns=["Archivo PDF", "Nombre Extraído"])
                ruta_salida = CARPETA_DESCARGAS / "nombres_extraidos.xlsx"
                df_nombres.to_excel(ruta_salida, index=False)
                
                self.ocultar_progreso()
                messagebox.showinfo("Éxito", f"✅ Nombres extraídos: {len(nombres_extraidos)}\n📁 Guardado en: {ruta_salida}")
            else:
                self.ocultar_progreso()
                messagebox.showwarning("Atención", "No se pudieron extraer nombres de ningún PDF.")
        
        threading.Thread(target=proceso, daemon=True).start()
    
    def comparar_nombres(self):
        """Compara nombres de PDFs con Excel."""
        carpeta_seleccionada = filedialog.askdirectory(title="Seleccionar carpeta de PDFs", initialdir=str(Path.home()))
        if not carpeta_seleccionada:
            return
        
        archivo_excel_seleccionado = filedialog.askopenfilename(
            title="Seleccionar archivo Excel de validación",
            filetypes=[("Archivos Excel", "*.xlsx *.xls")],
            initialdir=str(Path.home())
        )
        if not archivo_excel_seleccionado:
            return
        
        def proceso():
            self.mostrar_progreso("Comparando nombres...")
            
            # Extraer nombres de PDFs
            carpeta_pdfs_path = Path(carpeta_seleccionada)
            nombres_pdf = []
            
            for item in carpeta_pdfs_path.iterdir():
                if item.is_file() and item.suffix.lower() == ".pdf":
                    nombre = extraer_nombre_desde_pdf(item)
                    if nombre:
                        nombres_pdf.append(nombre)
            
            # Leer nombres de Excel
            try:
                df = pd.read_excel(archivo_excel_seleccionado, usecols=[1], skiprows=6, header=None)
                nombres_excel = df.iloc[:, 0].dropna().astype(str).tolist()
            except Exception as e:
                self.ocultar_progreso()
                messagebox.showerror("Error", f"Error al leer Excel: {e}")
                return
            
            # Comparar
            resultados = comparar_nombres_fuzzy(nombres_pdf, nombres_excel)
            
            # Exportar
            self.exportar_resultados_nombres(resultados)
            self.ocultar_progreso()
        
        threading.Thread(target=proceso, daemon=True).start()
    
    def extraer_documentos(self):
        """Extrae documentos de PDFs y los guarda en Excel."""
        carpeta_seleccionada = filedialog.askdirectory(title="Seleccionar carpeta de PDFs", initialdir=str(Path.home()))
        if not carpeta_seleccionada:
            return
        
        def proceso():
            self.mostrar_progreso("Extrayendo documentos de PDFs...")
            
            carpeta_pdfs_path = Path(carpeta_seleccionada)
            documentos_extraidos = []
            archivos_no_leidos = []
            
            for item in carpeta_pdfs_path.iterdir():
                if item.is_file() and item.suffix.lower() == ".pdf":
                    numero_doc, tipo_doc = extraer_documento_desde_pdf(item)
                    if numero_doc and tipo_doc:
                        documentos_extraidos.append([item.name, numero_doc, tipo_doc])
                    else:
                        archivos_no_leidos.append(item.name)
            
            if documentos_extraidos:
                df_documentos = pd.DataFrame(documentos_extraidos, columns=["Archivo PDF", "Número Documento", "Tipo Documento"])
                ruta_salida = CARPETA_DESCARGAS / "documentos_extraidos.xlsx"
                df_documentos.to_excel(ruta_salida, index=False)
                
                self.ocultar_progreso()
                messagebox.showinfo("Éxito", f"✅ Documentos extraídos: {len(documentos_extraidos)}\n📁 Guardado en: {ruta_salida}")
            else:
                self.ocultar_progreso()
                messagebox.showwarning("Atención", "No se pudieron extraer documentos de ningún PDF.")
        
        threading.Thread(target=proceso, daemon=True).start()
    
    def comparar_documentos(self):
        """Compara documentos de PDFs con Excel."""
        carpeta_seleccionada = filedialog.askdirectory(title="Seleccionar carpeta de PDFs", initialdir=str(Path.home()))
        if not carpeta_seleccionada:
            return
        
        archivo_excel_seleccionado = filedialog.askopenfilename(
            title="Seleccionar archivo Excel de validación",
            filetypes=[("Archivos Excel", "*.xlsx *.xls")],
            initialdir=str(Path.home())
        )
        if not archivo_excel_seleccionado:
            return
        
        def proceso():
            self.mostrar_progreso("Comparando documentos...")
            
            # Extraer documentos de PDFs
            carpeta_pdfs_path = Path(carpeta_seleccionada)
            documentos_pdf = []
            
            for item in carpeta_pdfs_path.iterdir():
                if item.is_file() and item.suffix.lower() == ".pdf":
                    numero_doc, tipo_doc = extraer_documento_desde_pdf(item)
                    if numero_doc and tipo_doc:
                        documentos_pdf.append((numero_doc, tipo_doc, item.name))
            
            # Leer documentos de Excel
            try:
                df = pd.read_excel(archivo_excel_seleccionado, usecols=[0], skiprows=6, header=None)
                documentos_raw = df.iloc[:, 0].dropna().astype(str).tolist()
                documentos_excel = [normalizar_documento(doc) for doc in documentos_raw]
            except Exception as e:
                self.ocultar_progreso()
                messagebox.showerror("Error", f"Error al leer Excel: {e}")
                return
            
            # Comparar
            resultados = comparar_documentos_exactos(documentos_pdf, documentos_excel)
            
            # Exportar
            self.exportar_resultados_documentos(resultados)
            self.ocultar_progreso()
        
        threading.Thread(target=proceso, daemon=True).start()
    
    def validacion_completa(self):
        """Realiza validación completa de nombres y documentos."""
        carpeta_seleccionada = filedialog.askdirectory(title="Seleccionar carpeta de PDFs", initialdir=str(Path.home()))
        if not carpeta_seleccionada:
            return
        
        archivo_excel_seleccionado = filedialog.askopenfilename(
            title="Seleccionar archivo Excel de validación",
            filetypes=[("Archivos Excel", "*.xlsx *.xls")],
            initialdir=str(Path.home())
        )
        if not archivo_excel_seleccionado:
            return
        
        def proceso():
            self.mostrar_progreso("Realizando validación completa...")
            
            carpeta_pdfs_path = Path(carpeta_seleccionada)
            resultados_completos = []
            
            # Leer datos de Excel
            try:
                df_nombres = pd.read_excel(archivo_excel_seleccionado, usecols=[1], skiprows=6, header=None)
                nombres_excel = df_nombres.iloc[:, 0].dropna().astype(str).tolist()
                
                df_docs = pd.read_excel(archivo_excel_seleccionado, usecols=[0], skiprows=6, header=None)
                documentos_raw = df_docs.iloc[:, 0].dropna().astype(str).tolist()
                documentos_excel = [normalizar_documento(doc) for doc in documentos_raw]
            except Exception as e:
                self.ocultar_progreso()
                messagebox.showerror("Error", f"Error al leer Excel: {e}")
                return
            
            # Procesar cada PDF
            for item in carpeta_pdfs_path.iterdir():
                if item.is_file() and item.suffix.lower() == ".pdf":
                    # Extraer nombre
                    nombre_extraido = extraer_nombre_desde_pdf(item)
                    
                    # Extraer documento
                    numero_doc, tipo_doc = extraer_documento_desde_pdf(item)
                    
                    # Validar nombre
                    estado_nombre = "❌ NO EXTRAÍDO"
                    mejor_nombre = ""
                    similitud_nombre = 0
                    
                    if nombre_extraido:
                        nombres_excel_normalizados = [normalizar_texto(n) for n in nombres_excel]
                        nombre_normalizado = normalizar_texto(nombre_extraido)
                        
                        try:
                            idx = nombres_excel_normalizados.index(nombre_normalizado)
                            mejor_nombre = nombres_excel[idx]
                            similitud_nombre = 100
                            estado_nombre = "✔ EXACTA"
                        except ValueError:
                            for i, nombre_excel_norm in enumerate(nombres_excel_normalizados):
                                if not nombre_excel_norm:
                                    continue
                                score = fuzz.ratio(nombre_normalizado, nombre_excel_norm)
                                if score > similitud_nombre:
                                    similitud_nombre = score
                                    mejor_nombre = nombres_excel[i]
                            
                            if similitud_nombre >= 90:
                                estado_nombre = f"⚠️ ALTA ({similitud_nombre}%)"
                            elif similitud_nombre >= 70:
                                estado_nombre = f"⚠️ MEDIA ({similitud_nombre}%)"
                            elif similitud_nombre >= 50:
                                estado_nombre = f"⚠️ BAJA ({similitud_nombre}%)"
                            else:
                                estado_nombre = "❌ SIN COINCIDENCIA"
                    
                    # Validar documento
                    estado_documento = "❌ NO EXTRAÍDO"
                    mejor_documento = ""
                    similitud_documento = 0
                    
                    if numero_doc:
                        if numero_doc in documentos_excel:
                            mejor_documento = numero_doc
                            similitud_documento = 100
                            estado_documento = "✔ EXACTA"
                        else:
                            for doc_excel in documentos_excel:
                                score = fuzz.ratio(numero_doc, doc_excel)
                                if score > similitud_documento:
                                    similitud_documento = score
                                    mejor_documento = doc_excel
                            
                            if similitud_documento >= 90:
                                estado_documento = f"⚠️ ALTA ({similitud_documento}%)"
                            elif similitud_documento >= 70:
                                estado_documento = f"⚠️ MEDIA ({similitud_documento}%)"
                            elif similitud_documento >= 50:
                                estado_documento = f"⚠️ BAJA ({similitud_documento}%)"
                            else:
                                estado_documento = "❌ SIN COINCIDENCIA"
                    
                    resultados_completos.append([
                        item.name,
                        nombre_extraido or "NO EXTRAÍDO",
                        mejor_nombre,
                        similitud_nombre,
                        estado_nombre,
                        numero_doc or "NO EXTRAÍDO",
                        tipo_doc or "NO IDENTIFICADO",
                        mejor_documento,
                        similitud_documento,
                        estado_documento
                    ])
            
            # Exportar resultados completos
            self.exportar_resultados_completos(resultados_completos)
            self.ocultar_progreso()
        
        threading.Thread(target=proceso, daemon=True).start()
    
    def exportar_resultados_nombres(self, resultados: List[List[Any]]):
        """Exporta resultados de comparación de nombres a Excel con colores."""
        df_resultados = pd.DataFrame(resultados, columns=["Nombre PDF", "Mejor Coincidencia Excel", "% Similitud", "Estado"])
        ruta_salida = CARPETA_DESCARGAS / "comparacion_nombres.xlsx"
        
        try:
            df_resultados.to_excel(ruta_salida, index=False)
            
            wb = load_workbook(ruta_salida)
            ws = wb.active
            
            # Aplicar formato de encabezados
            for col in range(1, 5):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Aplicar colores según estado
            for row_idx in range(2, ws.max_row + 1):
                estado_celda = ws[f"D{row_idx}"].value
                
                if "EXACTA" in str(estado_celda):
                    fill_color = VERDE
                elif "ALTA" in str(estado_celda) or "MEDIA" in str(estado_celda):
                    fill_color = AMARILLO
                elif "BAJA" in str(estado_celda):
                    fill_color = AZUL
                else:
                    fill_color = ROJO
                
                for col_letter in ["A", "B", "C", "D"]:
                    ws[f"{col_letter}{row_idx}"].fill = fill_color
            
            wb.save(ruta_salida)
            messagebox.showinfo("Éxito", f"✅ Comparación de nombres completada\n📁 Guardado en: {ruta_salida}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {e}")
    
    def exportar_resultados_documentos(self, resultados: List[List[Any]]):
        """Exporta resultados de comparación de documentos a Excel con colores."""
        df_resultados = pd.DataFrame(resultados, columns=[
            "Archivo PDF", "Documento Extraído", "Tipo Documento", 
            "Mejor Coincidencia Excel", "% Similitud", "Estado"
        ])
        ruta_salida = CARPETA_DESCARGAS / "comparacion_documentos.xlsx"
        
        try:
            df_resultados.to_excel(ruta_salida, index=False)
            
            wb = load_workbook(ruta_salida)
            ws = wb.active
            
            # Aplicar formato de encabezados
            for col in range(1, 7):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Aplicar colores según estado
            for row_idx in range(2, ws.max_row + 1):
                estado_celda = ws[f"F{row_idx}"].value
                
                if "EXACTA" in str(estado_celda):
                    fill_color = VERDE
                elif "ALTA" in str(estado_celda) or "MEDIA" in str(estado_celda):
                    fill_color = AMARILLO
                elif "BAJA" in str(estado_celda):
                    fill_color = AZUL
                else:
                    fill_color = ROJO
                
                for col_letter in ["A", "B", "C", "D", "E", "F"]:
                    ws[f"{col_letter}{row_idx}"].fill = fill_color
            
            wb.save(ruta_salida)
            messagebox.showinfo("Éxito", f"✅ Comparación de documentos completada\n📁 Guardado en: {ruta_salida}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {e}")
    
    def exportar_resultados_completos(self, resultados: List[List[Any]]):
        """Exporta resultados de validación completa a Excel con múltiples hojas."""
        ruta_salida = CARPETA_DESCARGAS / "validacion_completa.xlsx"
        
        try:
            # Crear DataFrame principal
            df_completo = pd.DataFrame(resultados, columns=[
                "Archivo PDF", "Nombre Extraído", "Mejor Nombre Excel", "% Similitud Nombre", "Estado Nombre",
                "Documento Extraído", "Tipo Documento", "Mejor Documento Excel", "% Similitud Documento", "Estado Documento"
            ])
            
            # Crear archivo Excel con múltiples hojas
            with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
                df_completo.to_excel(writer, sheet_name='Validación Completa', index=False)
                
                # Hoja de estadísticas
                stats_data = []
                total_archivos = len(resultados)
                
                # Estadísticas de nombres
                nombres_exactos = sum(1 for r in resultados if "EXACTA" in str(r[4]))
                nombres_similares = sum(1 for r in resultados if "ALTA" in str(r[4]) or "MEDIA" in str(r[4]))
                nombres_sin_coincidencia = sum(1 for r in resultados if "SIN COINCIDENCIA" in str(r[4]) or "NO EXTRAÍDO" in str(r[4]))
                
                # Estadísticas de documentos
                docs_exactos = sum(1 for r in resultados if "EXACTA" in str(r[9]))
                docs_similares = sum(1 for r in resultados if "ALTA" in str(r[9]) or "MEDIA" in str(r[9]))
                docs_sin_coincidencia = sum(1 for r in resultados if "SIN COINCIDENCIA" in str(r[9]) or "NO EXTRAÍDO" in str(r[9]))
                
                stats_data = [
                    ["ESTADÍSTICAS GENERALES", "", ""],
                    ["Total de archivos procesados", total_archivos, ""],
                    ["", "", ""],
                    ["VALIDACIÓN DE NOMBRES", "", ""],
                    ["Coincidencias exactas", nombres_exactos, f"{(nombres_exactos/total_archivos)*100:.1f}%"],
                    ["Similitudes parciales", nombres_similares, f"{(nombres_similares/total_archivos)*100:.1f}%"],
                    ["Sin coincidencia", nombres_sin_coincidencia, f"{(nombres_sin_coincidencia/total_archivos)*100:.1f}%"],
                    ["", "", ""],
                    ["VALIDACIÓN DE DOCUMENTOS", "", ""],
                    ["Coincidencias exactas", docs_exactos, f"{(docs_exactos/total_archivos)*100:.1f}%"],
                    ["Similitudes parciales", docs_similares, f"{(docs_similares/total_archivos)*100:.1f}%"],
                    ["Sin coincidencia", docs_sin_coincidencia, f"{(docs_sin_coincidencia/total_archivos)*100:.1f}%"]
                ]
                
                df_stats = pd.DataFrame(stats_data, columns=["Métrica", "Cantidad", "Porcentaje"])
                df_stats.to_excel(writer, sheet_name='Estadísticas', index=False)
            
            # Aplicar formato con openpyxl
            wb = load_workbook(ruta_salida)
            
            # Formatear hoja principal
            ws_main = wb['Validación Completa']
            for col in range(1, 11):
                cell = ws_main.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Aplicar colores según estados
            for row_idx in range(2, ws_main.max_row + 1):
                estado_nombre = ws_main[f"E{row_idx}"].value
                estado_documento = ws_main[f"J{row_idx}"].value
                
                # Color para nombres (columnas A-E)
                if "EXACTA" in str(estado_nombre):
                    fill_nombre = VERDE
                elif "ALTA" in str(estado_nombre) or "MEDIA" in str(estado_nombre):
                    fill_nombre = AMARILLO
                elif "BAJA" in str(estado_nombre):
                    fill_nombre = AZUL
                else:
                    fill_nombre = ROJO
                
                # Color para documentos (columnas F-J)
                if "EXACTA" in str(estado_documento):
                    fill_documento = VERDE
                elif "ALTA" in str(estado_documento) or "MEDIA" in str(estado_documento):
                    fill_documento = AMARILLO
                elif "BAJA" in str(estado_documento):
                    fill_documento = AZUL
                else:
                    fill_documento = ROJO
                
                # Aplicar colores
                for col_letter in ["A", "B", "C", "D", "E"]:
                    ws_main[f"{col_letter}{row_idx}"].fill = fill_nombre
                
                for col_letter in ["F", "G", "H", "I", "J"]:
                    ws_main[f"{col_letter}{row_idx}"].fill = fill_documento
            
            wb.save(ruta_salida)
            messagebox.showinfo("Éxito", f"✅ Validación completa finalizada\n📁 Guardado en: {ruta_salida}\n\n📊 Estadísticas incluidas en hoja separada")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {e}")
    
    def ejecutar(self):
        """Ejecuta la aplicación."""
        self.ventana.mainloop()

# --- Punto de Entrada del Script ---

if __name__ == "__main__":
    CARPETA_DESCARGAS.mkdir(parents=True, exist_ok=True)
    app = ValidadorUnificado()
    app.ejecutar()
