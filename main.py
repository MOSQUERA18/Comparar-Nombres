import re
import unicodedata
import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
from typing import List, Tuple, Optional, Any

import pandas as pd
import pdfplumber
from fuzzywuzzy import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# --- Constantes ---
CARPETA_DESCARGAS = Path.home() / "Downloads"
NOMBRE_ARCHIVO_SALIDA_COMPARACION = "comparacion_nombres.xlsx"
NOMBRE_ARCHIVO_SALIDA_EXTRACCION = "nombres_extraidos.xlsx"

# Patrones Regex para extracción de nombres
PATRON_GENERAL = re.compile(r"A nombre de:\s*(.+?)\s*Estado:", re.IGNORECASE)
PATRON_REGISTRO_CIVIL = re.compile(r"Registro Civil,\s*(.*?)\s*tiene inscrito", re.IGNORECASE | re.DOTALL)
PATRON_MIGRACION = re.compile(
    r"el migrante venezolano\s+([A-ZÁÉÍÓÚÑ ]{5,})\s+surtió",
    re.IGNORECASE | re.DOTALL
)

# Estilos para Excel
VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ROJO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# --- Funciones de Utilidad ---

def normalizar(texto: str) -> str:
    """
    Normaliza un texto: convierte a mayúsculas, elimina tildes, puntuación y espacios extra.
    """
    if not isinstance(texto, str):
        texto = str(texto)
        
    texto = texto.strip().upper()
    # Eliminar tildes (descomponer y mantener solo ASCII)
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('utf-8')
    # Eliminar puntuación (cualquier cosa que no sea letra A-Z o espacio)
    # Si se esperan números en los nombres, añadir 0-9 a la regex: [^A-Z0-9\s]
    texto = re.sub(r"[^A-Z\s]", "", texto)
    # Normalizar espacios (reemplazar múltiples espacios/tabulaciones/newlines con uno solo)
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()

def invertir_nombre(nombre: str) -> str:
    """
    Invierte un nombre si consta de múltiples partes, útil para formatos "APELLIDOS NOMBRES".
    Ejemplo: "PEREZ GONZALEZ JUAN CARLOS" -> "JUAN CARLOS PEREZ GONZALEZ"
    """
    partes = nombre.strip().split()
    if len(partes) >= 2:
        mitad = len(partes) // 2
        return " ".join(partes[mitad:] + partes[:mitad])
    return nombre

# --- Lógica Principal de Procesamiento ---

def extraer_nombres_desde_pdf(pdf_path: Path) -> Optional[str]:
    """
    Extrae un nombre de un archivo PDF utilizando una serie de patrones regex.
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            texto_completo = ""
            for pagina in pdf.pages:
                if pagina.extract_text(): # Asegurarse que hay texto
                    texto_completo += pagina.extract_text() + "\n"

            match_general = PATRON_GENERAL.search(texto_completo)
            if match_general:
                return match_general.group(1).strip()

            match_registro_civil = PATRON_REGISTRO_CIVIL.search(texto_completo)
            if match_registro_civil:
                # El nombre extraído podría estar en formato "APELLIDOS, NOMBRES"
                # Invertir para tener "NOMBRES APELLIDOS"
                return invertir_nombre(match_registro_civil.group(1).strip())

            match_migracion = PATRON_MIGRACION.search(texto_completo)
            if match_migracion:
                return match_migracion.group(1).strip()
                
    except Exception as e:
        print(f"Error al procesar el PDF {pdf_path.name}: {e}")
    return None

def leer_nombres_desde_excel(ruta_excel: Path) -> List[str]:
    """
    Lee nombres de una columna específica en un archivo Excel.
    Asume que los nombres están en la segunda columna (índice 1) y omite las primeras 6 filas.
    """
    try:
        # Quitado engine='xlrd' para que pandas use el motor adecuado (openpyxl para .xlsx)
        df = pd.read_excel(ruta_excel, usecols=[1], skiprows=6, header=None)
        return df.iloc[:, 0].dropna().astype(str).tolist()
    except Exception as e:
        messagebox.showerror("Error de lectura Excel", f"No se pudo leer el archivo Excel:\n{e}")
        return []

def _procesar_carpeta_pdfs(carpeta_path: Path) -> Tuple[List[str], List[str]]:
    """
    Procesa todos los archivos PDF en una carpeta, extrayendo nombres.
    Helper function.

    Returns:
        Una tupla (nombres_extraidos, archivos_no_leidos).
    """
    nombres_extraidos: List[str] = []
    archivos_no_leidos: List[str] = []

    for item in carpeta_path.iterdir():
        if item.is_file() and item.suffix.lower() == ".pdf":
            nombre = extraer_nombres_desde_pdf(item)
            if nombre:
                nombres_extraidos.append(nombre)
            else:
                archivos_no_leidos.append(item.name)
    return nombres_extraidos, archivos_no_leidos

def comparar_nombres_pdf_excel(nombres_pdf: List[str], nombres_excel: List[str]) -> List[List[Any]]:
    """
    Compara una lista de nombres extraídos de PDFs con una lista de nombres de un Excel.
    Utiliza fuzzy matching para calcular la similitud.
    """
    resultados: List[List[Any]] = []
    nombres_excel_normalizados = [normalizar(n) for n in nombres_excel]

    for nombre_pdf_original in nombres_pdf:
        nombre_pdf_normalizado = normalizar(nombre_pdf_original)
        
        if not nombre_pdf_normalizado: # Si la normalización resulta en string vacío
            resultados.append([nombre_pdf_original, "NOMBRE PDF VACIO/INVALIDO", 0, "❌"])
            continue

        mejor_match_excel_original = ""
        mejor_score = 0
        
        # Si se busca una coincidencia exacta primero (más rápido)
        try:
            idx = nombres_excel_normalizados.index(nombre_pdf_normalizado)
            mejor_match_excel_original = nombres_excel[idx]
            mejor_score = 100
        except ValueError:
            # Si no hay coincidencia exacta, usar fuzzy matching
            for i, nombre_excel_norm in enumerate(nombres_excel_normalizados):
                if not nombre_excel_norm: # Saltar nombres de Excel vacíos/inválidos
                    continue
                score = fuzz.ratio(nombre_pdf_normalizado, nombre_excel_norm)
                if score > mejor_score:
                    mejor_score = score
                    mejor_match_excel_original = nombres_excel[i]
        
        # Definir estado basado en el score (ej. 100% para ✔, podría ajustarse)
        estado = "✔" if mejor_score == 100 else "❌"
        if mejor_score > 0 and mejor_score < 100 : # Marcar como parcial si hay similitud pero no es exacta
             estado = f"⚠️ ({mejor_score}%)" # O alguna otra indicación



        resultados.append([nombre_pdf_original, mejor_match_excel_original, mejor_score, estado])
        
    return resultados

def exportar_resultados_a_excel_con_colores(resultados: List[List[Any]], ruta_salida_completa: Path) -> None:
    """
    Exporta los resultados de la comparación a un archivo Excel, coloreando las filas.
    """
    df_resultados = pd.DataFrame(resultados, columns=["Nombre PDF", "Nombre Excel Coincidente", "Similitud (%)", "Estado"])
    
    try:
        df_resultados.to_excel(ruta_salida_completa, index=False)
        
        wb = load_workbook(ruta_salida_completa)
        ws = wb.active
        
        # Aplicar formato de color
        for row_idx in range(2, ws.max_row + 1): # Filas de datos comienzan en 2
            # El estado está en la columna D (índice 3 en 0-based, 4 en 1-based)
            estado_celda = ws[f"D{row_idx}"].value 
            fill_color = VERDE if estado_celda == "✔" else (ROJO if estado_celda == "❌" else None) # No colorear si es parcial o error
            
            if "⚠️" in str(estado_celda): # Si es una advertencia (parcial)
                 # Podrías usar otro color, ej. amarillo
                 AMARILLO = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                 fill_color = AMARILLO

            if fill_color:
                for col_letter in ["A", "B", "C", "D"]:
                    ws[f"{col_letter}{row_idx}"].fill = fill_color
            
        wb.save(ruta_salida_completa)
        messagebox.showinfo("Exportación Exitosa", f"Comparación exportada y formateada en:\n{ruta_salida_completa}")
    except Exception as e:
        messagebox.showerror("Error de Exportación", f"No se pudo exportar el archivo Excel:\n{e}")


# --- Funciones de Interfaz de Usuario (Callbacks) ---

def _accion_extraer_nombres() -> None:
    """
    Callback para el botón "Extraer nombres de PDFs".
    Pide al usuario una carpeta, procesa los PDFs y guarda los nombres en un Excel.
    """
    carpeta_seleccionada = filedialog.askdirectory(title="Seleccionar carpeta de PDFs", initialdir=str(Path.home()))
    if not carpeta_seleccionada:
        return

    carpeta_pdfs_path = Path(carpeta_seleccionada)
    nombres_extraidos, no_leidos = _procesar_carpeta_pdfs(carpeta_pdfs_path)

    if not nombres_extraidos and not no_leidos:
        messagebox.showinfo("Resultado", "No se encontraron archivos PDF en la carpeta seleccionada.")
        return
    
    if nombres_extraidos:
        nombres_ordenados = sorted(nombres_extraidos, key=lambda n: normalizar(n))
        df_nombres = pd.DataFrame(nombres_ordenados, columns=["Nombres Extraídos"])
        
        ruta_salida_extraccion = CARPETA_DESCARGAS / NOMBRE_ARCHIVO_SALIDA_EXTRACCION
        try:
            df_nombres.to_excel(ruta_salida_extraccion, index=False)
            msg = f"Nombres extraídos y guardados en:\n{ruta_salida_extraccion}"
        except Exception as e:
            msg = f"Error al guardar nombres extraídos:\n{e}"
            messagebox.showerror("Error al Guardar", msg)
            return
    else:
        msg = "No se pudieron extraer nombres de ningún PDF."

    if no_leidos:
        msg += f"\n\nArchivos PDF no procesados o sin nombre encontrado:\n- " + "\n- ".join(no_leidos)
    
    messagebox.showinfo("Resultado de Extracción", msg)

def _accion_comparar_nombres() -> None:
    """
    Callback para el botón "Comparar con Excel".
    Pide carpeta de PDFs y archivo Excel, compara nombres y exporta resultados.
    """
    carpeta_seleccionada = filedialog.askdirectory(title="Seleccionar carpeta de PDFs", initialdir=str(Path.home()))
    if not carpeta_seleccionada:
        return
    
    archivo_excel_seleccionado = filedialog.askopenfilename(
        title="Seleccionar archivo Excel de validación",
        filetypes=[("Archivos Excel", "*.xlsx *.xls")],
        initialdir=str(Path.home())
    )
    if not archivo_excel_seleccionado:
        return

    carpeta_pdfs_path = Path(carpeta_seleccionada)
    excel_validacion_path = Path(archivo_excel_seleccionado)

    nombres_pdf, no_leidos_pdf = _procesar_carpeta_pdfs(carpeta_pdfs_path)
    
    if not nombres_pdf:
        messagebox.showwarning("Atención", "No se extrajeron nombres de ningún PDF en la carpeta seleccionada.")
        if no_leidos_pdf:
             messagebox.showinfo("PDFs no leídos", "Archivos PDF no procesados o sin nombre encontrado:\n\n" + "\n".join(no_leidos_pdf))
        return

    nombres_excel = leer_nombres_desde_excel(excel_validacion_path)
    if not nombres_excel:
        messagebox.showwarning("Atención", "No se encontraron nombres en el archivo Excel o hubo un error al leerlo.")
        return

    resultados_comparacion = comparar_nombres_pdf_excel(nombres_pdf, nombres_excel)
    
    ruta_salida_comparacion = CARPETA_DESCARGAS / NOMBRE_ARCHIVO_SALIDA_COMPARACION
    exportar_resultados_a_excel_con_colores(resultados_comparacion, ruta_salida_comparacion)

    if no_leidos_pdf:
        messagebox.showwarning("Atención - PDFs no leídos", 
                               "No se pudo extraer nombre de los siguientes archivos PDF:\n\n" + "\n".join(no_leidos_pdf))

# --- Configuración y Lanzamiento de la GUI ---

def lanzar_gui() -> None:
    """
    Crea y lanza la interfaz gráfica de usuario (GUI) con Tkinter.
    """
    ventana = tk.Tk()
    ventana.title("Validador de Nombres PDF vs Excel")
    ventana.geometry("450x250") # Ajustado tamaño para mejor visualización
    
    # Estilo simple para los botones
    estilo_boton = {'relief': tk.RAISED, 'borderwidth': 2, 'font': ('Arial', 10, 'bold'), 'height': 2, 'width': 35, 'pady': 5}

    label_titulo = tk.Label(ventana, text="Herramienta de Validación de Nombres", font=('Arial', 14, 'bold'))
    label_titulo.pack(pady=(10,15))

    boton_extraer = tk.Button(ventana, text="Extraer Nombres de PDFs a Excel", command=_accion_extraer_nombres, **estilo_boton)
    boton_extraer.pack(pady=10)

    boton_comparar = tk.Button(ventana, text="Comparar Nombres de PDFs con Excel", command=_accion_comparar_nombres, **estilo_boton)
    boton_comparar.pack(pady=10)

    ventana.mainloop()

# --- Punto de Entrada del Script ---

if __name__ == "__main__":
    # Asegurarse que la carpeta de descargas exista, si no, crearla.
    CARPETA_DESCARGAS.mkdir(parents=True, exist_ok=True) 
    lanzar_gui()